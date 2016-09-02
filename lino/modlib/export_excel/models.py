# -*- coding: UTF-8 -*-
# Copyright 2014-2016 Josef Kejzlar, Luc Saffre, Hamza Khchine
# License: BSD (see file COPYING for details)

"""Database models for `lino.modlib.export_excel`.

"""
from builtins import str
import os

from django.conf import settings
from django.db.models import Model
from django.utils.functional import Promise
from lino.core import actions
from lino.core.tables import AbstractTable
from lino.utils.media import TmpMediaFile
from lino.utils.xmlgen.html import E
from lino.utils.quantities import Duration
from lino.core.choicelists import Choice
from lino.api import dd

from django.utils.translation import ugettext_lazy as _


def sheet_name(s):
    s = s[:31]
    for c in u"[]:\\?/*\x00":
        s = s.replace(c, '_')
    return s


def ar2workbook(ar, column_names=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    # local import to avoid the following traceback:
    # Error in sys.exitfunc:
    # Traceback (most recent call last):
    #   File "/usr/lib/python2.7/atexit.py", line 24, in _run_exitfuncs
    #     func(*targs, **kargs)
    #   File "/openpyxl/writer/write_only.py", line 38, in _openpyxl_shutdown
    #     for path in ALL_TEMP_FILES:
    # TypeError: 'NoneType' object is not iterable

    workbook = Workbook(guess_types=True)
    sheet = workbook.active
    sheet.title = sheet_name(ar.get_title())

    bold_font = Font(name='Calibri', size=11, bold=True, )

    fields, headers, widths = ar.get_field_info(column_names)

    for c, column in enumerate(fields):
        sheet.cell(row=1, column=c + 1).value = str(headers[c])
        sheet.cell(row=1, column=c + 1).font = bold_font
        # sheet.col(c).width = min(256 * widths[c] / 7, 65535)
        # 256 == 1 character width, max width=65535

    for c, column in enumerate(fields):
        for r, row in enumerate(ar.data_iterator, start=1):
            sf = column.field._lino_atomizer
            value = sf.full_value_from_object(row, ar)
            if type(value) == bool:
                value = value and 1 or 0
            elif isinstance(value, (Duration, Choice)):
                value = str(value)
            elif E.iselement(value):
                value = E.to_rst(value)
                # dd.logger.info("20160716 %s", value)
            elif isinstance(value, Promise):
                value = str(value)
            elif isinstance(value, Model):
                value = str(value)
            sheet.cell(row=r + 1, column=c + 1).value = value

    return workbook


class ExportExcelAction(actions.Action):
    label = _("Export to .xls")
    help_text = _('Export this table as an .xls document')
    icon_name = 'page_excel'
    sort_index = -5
    select_rows = False
    default_format = 'ajax'
    preprocessor = "Lino.get_current_grid_config"

    def is_callable_from(self, caller):
        return isinstance(caller, actions.GridEdit)

    def run_from_ui(self, ar, **kw):
        # Prepare tmp file
        mf = TmpMediaFile(ar, 'xlsx')
        settings.SITE.makedirs_if_missing(os.path.dirname(mf.name))

        # Render
        self.render(ar, mf.name)

        # Tell client that the action was successful and that it
        # should open a new browser window on the generated file.
        ar.success(open_url=mf.url)

    def render(self, ar, file):
        workbook = ar2workbook(ar)
        workbook.save(file)


AbstractTable.export_excel = ExportExcelAction()
